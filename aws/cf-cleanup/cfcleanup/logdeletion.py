"""Record a completed stack deletion in a batch's cleanup workbook.

Sets the tool-owned Deleted? column and appends a deletion stamp to Notes on the
matching 'Cleanup' row:
    Deleted?  ->  Yes
    Notes     ->  existing text + "Deleted YYYY-MM-DD" (appended, never overwritten)

Notes is a free-form, human-editable column; any reviewer text is preserved.

Run as:
    python3 -m cfcleanup log <stack-name> [--batch NAME|DIR] [--date YYYY-MM-DD]
"""
from __future__ import annotations

import argparse
import sys
from datetime import date

from openpyxl import load_workbook

from . import common as cf

SHEET = "Cleanup"


def col_index(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=1, column=c).value or "").strip().lower() == header.lower():
            return c
    raise SystemExit(f"ERROR: column '{header}' not found on '{ws.title}' sheet.")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="python3 -m cfcleanup log")
    ap.add_argument("stack")
    ap.add_argument("--batch")
    ap.add_argument("--date", default=date.today().isoformat())
    ap.add_argument("--workbook", help="Path to the reviewed workbook (defaults to CF_WORKBOOK, then the batch's own).")
    args = ap.parse_args(argv)

    batch = cf.resolve_batch(args.batch)
    wb_path = cf.resolve_workbook(batch, args.workbook)
    if wb_path is None:
        sys.exit(f"ERROR: no review workbook (*-cf-cleanup.xlsx) in batch {batch}")

    wb = load_workbook(wb_path)
    ws = wb[SHEET]
    stack_col = col_index(ws, "Stack")
    deleted_col = col_index(ws, "Deleted?")
    notes_col = col_index(ws, "Notes")

    for row in range(2, ws.max_row + 1):
        if (ws.cell(row=row, column=stack_col).value or "").strip() == args.stack:
            ws.cell(row=row, column=deleted_col).value = "Yes"
            notes_cell = ws.cell(row=row, column=notes_col)
            existing = (notes_cell.value or "").strip()
            stamp = f"Deleted {args.date}"
            notes_cell.value = f"{existing}; {stamp}" if existing else stamp
            wb.save(wb_path)
            print(f"Marked '{args.stack}' deleted ({args.date}) on row {row}.")
            return 0

    print(f"ERROR: stack '{args.stack}' not found on '{SHEET}' sheet.")
    return 1
