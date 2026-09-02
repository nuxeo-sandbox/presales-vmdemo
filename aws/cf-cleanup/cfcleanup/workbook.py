"""Build the collaborative Excel workbook for a cleanup batch.

Writes <batch>/<date>-cf-cleanup.xlsx. With no --batch, uses the most recent batch.

Columns (all human-owned):
  Yellow, editable: Decision, Deleted?, Notes

The tool never writes this workbook. The log command reads it read-only and
prints the Deleted?/Notes values for a human to enter by hand.

Run as:
    python3 -m cfcleanup workbook [--batch NAME|DIR]
"""
from __future__ import annotations

import argparse
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import common as cf

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # human-editable columns
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

AGE_FILLS = {
    "stale": PatternFill("solid", fgColor="F4CCCC"),
    "aging": PatternFill("solid", fgColor="FCE5CD"),
    "recent": PatternFill("solid", fgColor="D9EAD3"),
}


def age_bucket(age: float | None) -> str:
    if age is None:
        return "recent"
    if age >= 12:
        return "stale"
    if age >= 6:
        return "aging"
    return "recent"


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="python3 -m cfcleanup workbook")
    ap.add_argument("--batch", help="batch name or dir (default: most recent)")
    args = ap.parse_args(argv)

    batch = cf.resolve_batch(args.batch)
    meta = cf.load_meta(batch)
    rows, now = cf.load_rows(batch)
    demo, infra = cf.split_rows(rows)
    account = meta.get("account", "(unknown account)")

    wb = Workbook()

    # ============================ Sheet: Cleanup ===========================
    ws = wb.active
    ws.title = "Cleanup"
    headers = [
        "Region", "Stack", "Customer", "Owner", "Manager", "Studio project",
        "Nuxeo", "Last activity", "Age (mo)", "Class",
        "Decision", "Deleted?", "Notes",
    ]
    ws.append(headers)
    for r in demo:
        ws.append(
            [
                r["region"], r["name"], r["customer"], r["owner"], r["manager"],
                r["studio"], r["nuxeo"],
                r["last_activity"].date().isoformat() if r["last_activity"] else "",
                round(r["age_months"], 0) if r["age_months"] is not None else "",
                r["class"],
                "",  # Decision (human)
                "",  # Deleted? (human)
                "",  # Notes (free-form; human enters the deletion stamp)
            ]
        )

    ncols = len(headers)
    nrows = len(demo)
    style_header(ws, ncols)

    widths = [11.5, 30.3, 24.0, 15.8, 21.5, 26.0, 8.0, 13.0, 8.0, 27.0, 20.2, 9.8, 53.5]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    age_col, decision_col, deleted_col, notes_col = 9, 11, 12, 13
    for ri in range(2, nrows + 2):
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            cell.alignment = CENTER if ci in (1, 7, 9, 12) else LEFT
        r = demo[ri - 2]
        ws.cell(row=ri, column=age_col).fill = AGE_FILLS[age_bucket(r["age_months"])]
        # Decision, Deleted? and Notes are human-editable (yellow).
        for ci in (decision_col, deleted_col, notes_col):
            ws.cell(row=ri, column=ci).fill = INPUT_FILL

    decision_dv = DataValidation(
        type="list",
        formula1='"Keep - active engagement,Keep - generic demo,Delete,Investigate"',
        allow_blank=True,
    )
    decision_dv.prompt = "Choose a disposition for this stack"
    decision_dv.promptTitle = "Decision"
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{nrows + 1}")

    deleted_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(deleted_dv)
    deleted_dv.add(f"{get_column_letter(deleted_col)}2:{get_column_letter(deleted_col)}{nrows + 1}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"

    # ======================= Sheet: Excluded infra =========================
    ws2 = wb.create_sheet("Excluded infra")
    ws2.append(["Stack", "Regions", "Count"])
    infra_by_name: dict[str, list[str]] = defaultdict(list)
    for r in infra:
        infra_by_name[r["name"]].append(r["region"])

    def infra_sort_key(item: tuple[str, list[str]]) -> tuple[int, str]:
        name = item[0]
        gov = 1 if (name.startswith("StackSet-") or name.startswith("AWS-QuickSetup-")) else 0
        return (gov, name.lower())

    for name, regs in sorted(infra_by_name.items(), key=infra_sort_key):
        rs = ", ".join(
            f"{reg}\u00d7{regs.count(reg)}" if regs.count(reg) > 1 else reg
            for reg in cf.sort_regions(regs)
        )
        ws2.append([name, rs, len(regs)])
    style_header(ws2, 3)
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 8
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:C{len(infra_by_name) + 1}"

    # ============================ Sheet: README ============================
    ws3 = wb.create_sheet("README", 0)
    readme = [
        ("CloudFormation Stack Cleanup", True),
        (f"Account {account}  |  generated {now.date()}", False),
        ("", False),
        (f"{nrows} Nuxeo demo/customer stacks are listed on the 'Cleanup' tab.", False),
        ("Nested (NestedStackNEV-*) stacks are omitted \u2014 deleted with their parent.", False),
        (f"{len(infra_by_name)} platform/automation stacks are on 'Excluded infra'", False),
        ("(kept in place, not for deletion).", False),
        ("", False),
        ("How to use:", True),
        ("\u2022 Fill the yellow column: Decision.", False),
        ("\u2022 Decision options: Keep - active engagement / Keep - generic demo /", False),
        ("  Delete / Investigate (dropdown).", False),
        ("\u2022 Deleted?: set it to Yes yourself once the tooling reports DELETE_COMPLETE.", False),
        ("\u2022 Notes is free-form: add your own remarks and the \"Deleted YYYY-MM-DD\"", False),
        ("  stamp the tooling prints for you when a stack is removed.", False),
        ("\u2022 Use the header filters to sort by Region, Age, Class, or Owner.", False),
        ("\u2022 Age colouring: red >12mo, orange 6-12mo, green <6mo.", False),
        ("\u2022 Default is deletion \u2014 anything left without a 'Keep' will be removed.", False),
    ]
    for i, (text, bold) in enumerate(readme, start=1):
        c = ws3.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
    ws3.column_dimensions["A"].width = 72
    ws3.sheet_view.showGridLines = False

    out = cf.workbook_path(batch)
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  Cleanup sheet: {nrows} stacks")
    print(f"  Excluded infra sheet: {len(infra_by_name)} distinct ({len(infra)} total)")
    return 0
