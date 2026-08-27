"""Safety gate: report the human-owned Decision for a stack in a batch workbook.

Decision is owned by humans; deletion tooling must never write it and must
refuse to delete a stack a human has not marked "Delete".

Exit codes:
    0  -> Decision == "Delete"        (cleared to proceed)
    2  -> some other/blank Decision   (NOT cleared)
    1  -> stack not found on the sheet / usage error

Run as:
    python3 -m cfcleanup check <stack-name> [--batch NAME|DIR]
"""
from __future__ import annotations

import argparse
import sys

from openpyxl import load_workbook

from . import common as cf

SHEET = "Cleanup"


def col_index(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=1, column=c).value or "").strip().lower() == header.lower():
            return c
    print(f"ERROR: column '{header}' not found on '{ws.title}' sheet.", file=sys.stderr)
    sys.exit(1)


def check(stack: str, batch: str, workbook: str | None = None) -> int:
    """Return 0 if the stack's Decision is 'Delete', 2 if other/blank, 1 if absent."""
    wb_path = cf.resolve_workbook(batch, workbook)
    if wb_path is None:
        print(f"ERROR: no review workbook (*-cf-cleanup.xlsx) in batch {batch}", file=sys.stderr)
        return 1

    wb = load_workbook(wb_path, read_only=True)
    ws = wb[SHEET]
    stack_col = col_index(ws, "Stack")
    decision_col = col_index(ws, "Decision")

    for row in ws.iter_rows(min_row=2):
        if (row[stack_col - 1].value or "").strip() == stack:
            decision = (row[decision_col - 1].value or "").strip()
            if decision.lower() == "delete":
                print(f"OK: '{stack}' Decision = 'Delete'.")
                return 0
            print(f"BLOCKED: '{stack}' Decision = '{decision or '<blank>'}' (not 'Delete').")
            return 2

    print(f"NOT_FOUND: '{stack}' is not on the '{SHEET}' sheet.")
    return 1


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="python3 -m cfcleanup check")
    ap.add_argument("stack")
    ap.add_argument("--batch")
    ap.add_argument("--workbook", help="Path to the reviewed workbook (defaults to CF_WORKBOOK, then the batch's own).")
    args = ap.parse_args(argv)
    return check(args.stack, cf.resolve_batch(args.batch), args.workbook)
